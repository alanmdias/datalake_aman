"""
Gera o Data Catalog completo do schema sm.

Para cada tabela extrai:
  - Schema, nome, número exato de linhas, tamanho em disco
  - Número de colunas, chave primária, número de FKs
  - Última atualização (quando existir coluna de data)
  - Comentários da tabela e das colunas (quando existirem)

Saídas:
  - data_catalog.xlsx  (aba Resumo + aba Colunas)
  - docs/data_catalog.md
"""
import os
import io
import warnings
from dotenv import load_dotenv
import psycopg2
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()
warnings.filterwarnings("ignore")

SCHEMA = "sm"


# ── Queries ────────────────────────────────────────────────────────────────────

QUERY_TABLES = """
SELECT
    t.table_name,
    pg_size_pretty(pg_total_relation_size(quote_ident(t.table_schema) || '.' || quote_ident(t.table_name))) AS tamanho_disco,
    pg_total_relation_size(quote_ident(t.table_schema) || '.' || quote_ident(t.table_name))                 AS tamanho_bytes,
    obj_description(
        (quote_ident(t.table_schema) || '.' || quote_ident(t.table_name))::regclass
    ) AS comentario_tabela
FROM information_schema.tables t
WHERE t.table_schema = %s AND t.table_type = 'BASE TABLE'
ORDER BY t.table_name;
"""

QUERY_COLUMNS = """
SELECT
    c.table_name,
    c.column_name,
    c.ordinal_position,
    c.data_type,
    c.character_maximum_length,
    c.is_nullable,
    c.column_default,
    col_description(
        (quote_ident(c.table_schema) || '.' || quote_ident(c.table_name))::regclass,
        c.ordinal_position
    ) AS comentario_coluna
FROM information_schema.columns c
WHERE c.table_schema = %s
ORDER BY c.table_name, c.ordinal_position;
"""

QUERY_PKS = """
SELECT
    kcu.table_name,
    string_agg(kcu.column_name, ', ' ORDER BY kcu.ordinal_position) AS chave_primaria
FROM information_schema.table_constraints tc
JOIN information_schema.key_column_usage kcu
    ON tc.constraint_name = kcu.constraint_name
    AND tc.table_schema   = kcu.table_schema
WHERE tc.table_schema    = %s
  AND tc.constraint_type = 'PRIMARY KEY'
GROUP BY kcu.table_name;
"""

QUERY_FKS = """
SELECT
    tc.table_name,
    COUNT(*) AS num_fks,
    string_agg(
        kcu.column_name || ' -> ' || ccu.table_name || '.' || ccu.column_name,
        ' | ' ORDER BY kcu.column_name
    ) AS fks_detalhe
FROM information_schema.table_constraints tc
JOIN information_schema.key_column_usage kcu
    ON tc.constraint_name = kcu.constraint_name
    AND tc.table_schema   = kcu.table_schema
JOIN information_schema.constraint_column_usage ccu
    ON ccu.constraint_name = tc.constraint_name
    AND ccu.table_schema   = tc.table_schema
WHERE tc.table_schema    = %s
  AND tc.constraint_type = 'FOREIGN KEY'
GROUP BY tc.table_name;
"""

# Colunas candidatas a "última atualização"
DATE_COLUMN_CANDIDATES = [
    "updated_at", "updatedAt", "data_atualizacao", "dt_atualizacao",
    "data_alteracao", "dt_alteracao", "modified_at", "data_modificacao",
    "dt_modificacao", "data_update", "dt_update", "last_update",
    "dtalteracao", "dtatualizacao",
]


def get_last_update(conn, schema, table, columns):
    """Tenta obter o MAX de uma coluna de data de modificação."""
    col_names_lower = {c["column_name"].lower(): c["column_name"] for c in columns}
    for candidate in DATE_COLUMN_CANDIDATES:
        if candidate.lower() in col_names_lower:
            real_col = col_names_lower[candidate.lower()]
            try:
                with conn.cursor() as cur:
                    cur.execute(
                        f'SELECT MAX("{real_col}") FROM "{schema}"."{table}"'
                    )
                    result = cur.fetchone()[0]
                    return (str(result)[:10] if result else None, real_col)
            except Exception:
                conn.rollback()
    return (None, None)


def style_excel(ws, header_color="1F4E79"):
    header_fill = PatternFill(fill_type="solid", fgColor=header_color)
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 60)


def main():
    print("Conectando ao PostgreSQL...")
    conn = psycopg2.connect(
        host=os.getenv("PG_HOST"), port=os.getenv("PG_PORT"),
        dbname=os.getenv("PG_DB"), user=os.getenv("PG_USER"),
        password=os.getenv("PG_PASSWORD"), connect_timeout=10,
    )
    cur = conn.cursor()

    # ── Metadados base ─────────────────────────────────────────────────────────
    print("Coletando metadados das tabelas...")
    cur.execute(QUERY_TABLES, (SCHEMA,))
    tables_meta = {r[0]: {"tamanho_disco": r[1], "tamanho_bytes": r[2], "comentario": r[3]}
                   for r in cur.fetchall()}

    print("Coletando colunas...")
    cur.execute(QUERY_COLUMNS, (SCHEMA,))
    cols_raw = cur.fetchall()
    col_headers = ["table_name", "column_name", "ordinal_position", "data_type",
                   "character_maximum_length", "is_nullable", "column_default", "comentario_coluna"]
    columns_by_table = {}
    for row in cols_raw:
        rec = dict(zip(col_headers, row))
        columns_by_table.setdefault(rec["table_name"], []).append(rec)

    print("Coletando chaves primárias...")
    cur.execute(QUERY_PKS, (SCHEMA,))
    pks = {r[0]: r[1] for r in cur.fetchall()}

    print("Coletando chaves estrangeiras...")
    cur.execute(QUERY_FKS, (SCHEMA,))
    fks = {r[0]: {"num_fks": r[1], "detalhe": r[2]} for r in cur.fetchall()}

    # ── Contagem de linhas e última atualização (por tabela) ───────────────────
    table_names = list(tables_meta.keys())
    summary_rows = []
    total = len(table_names)

    for i, table in enumerate(table_names, 1):
        print(f"  [{i}/{total}] {table}...", end="\r")

        # Row count
        cur.execute(f'SELECT COUNT(*) FROM "{SCHEMA}"."{table}"')
        row_count = cur.fetchone()[0]

        # Última atualização
        cols = columns_by_table.get(table, [])
        last_upd, upd_col = get_last_update(conn, SCHEMA, table, cols)

        meta = tables_meta[table]
        pk = pks.get(table, "—")
        fk_info = fks.get(table, {"num_fks": 0, "detalhe": "—"})

        summary_rows.append({
            "Schema":            SCHEMA,
            "Tabela":            table,
            "Linhas":            row_count,
            "Tamanho em disco":  meta["tamanho_disco"],
            "Colunas":           len(cols),
            "Chave primária":    pk,
            "Nº de FKs":         fk_info["num_fks"],
            "FKs (detalhe)":     fk_info["detalhe"],
            "Última atualização": last_upd,
            "Coluna de data":    upd_col,
            "Comentário tabela": meta["comentario"],
        })

    print(f"\nColetado {total} tabelas.")
    conn.close()

    # ── DataFrame de resumo ────────────────────────────────────────────────────
    df_summary = pd.DataFrame(summary_rows)

    # ── DataFrame de colunas ──────────────────────────────────────────────────
    col_rows = []
    for table, cols in columns_by_table.items():
        for c in cols:
            col_rows.append({
                "Schema":      SCHEMA,
                "Tabela":      table,
                "Coluna":      c["column_name"],
                "Posição":     c["ordinal_position"],
                "Tipo":        c["data_type"],
                "Tamanho máx": c["character_maximum_length"],
                "Nullable":    c["is_nullable"],
                "Default":     c["column_default"],
                "Comentário":  c["comentario_coluna"],
            })
    df_columns = pd.DataFrame(col_rows)

    # ── Excel ──────────────────────────────────────────────────────────────────
    output_xlsx = "data_catalog.xlsx"
    print(f"Gerando {output_xlsx}...")
    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        df_summary.to_excel(writer, index=False, sheet_name="Resumo")
        df_columns.to_excel(writer, index=False, sheet_name="Colunas")

        style_excel(writer.sheets["Resumo"])
        style_excel(writer.sheets["Colunas"])

        # Cor por volume na aba Resumo
        ws = writer.sheets["Resumo"]
        for row in ws.iter_rows(min_row=2):
            linhas = row[2].value or 0
            if linhas == 0:
                color = "F4CCCC"   # vermelho claro — vazia
            elif linhas > 900_000:
                color = "FFEB9C"   # amarelo — grande
            else:
                color = "C6EFCE"   # verde — normal
            for cell in row:
                cell.fill = PatternFill(fill_type="solid", fgColor=color)

    # ── Markdown ───────────────────────────────────────────────────────────────
    output_md = "docs/data_catalog.md"
    print(f"Gerando {output_md}...")

    md_lines = [
        "# Data Catalog — San Marino (schema sm)",
        "",
        f"Base de dados: `sm_teste` | Schema: `sm` | Total de tabelas: {total}",
        "",
        "> Gerado automaticamente por `etapa1_data_catalog.py`.",
        "> Tamanhos incluem índices. Última atualização detectada via colunas de data de modificação (quando existentes).",
        "",
        "## Resumo por tabela",
        "",
        "| Schema | Tabela | Linhas | Tamanho | Colunas | Chave Primária | Nº FKs | Última Atualização |",
        "| ------ | ------ | -----: | ------- | ------: | -------------- | -----: | ------------------ |",
    ]

    for r in summary_rows:
        linhas_fmt = f"{r['Linhas']:,}".replace(",", ".")
        pk = r["Chave primária"] or "—"
        last = r["Última atualização"] or "—"
        md_lines.append(
            f"| {r['Schema']} | {r['Tabela']} | {linhas_fmt} | {r['Tamanho em disco']} "
            f"| {r['Colunas']} | {pk} | {r['Nº de FKs']} | {last} |"
        )

    md_lines += [
        "",
        "---",
        "",
        "## Legenda",
        "",
        "- **Linhas**: contagem exata via `COUNT(*)`",
        "- **Tamanho**: `pg_total_relation_size` (dados + índices + TOAST)",
        "- **Chave Primária**: coluna(s) que compõem a PK, conforme `information_schema`",
        "- **Nº FKs**: quantidade de foreign keys declaradas na tabela",
        "- **Última Atualização**: MAX de coluna de data de modificação quando identificada",
        "",
        "## Tabelas com maior volume",
        "",
        "| Tabela | Linhas | Tamanho |",
        "| ------ | -----: | ------- |",
    ]

    top10 = sorted(summary_rows, key=lambda x: x["Linhas"], reverse=True)[:10]
    for r in top10:
        linhas_fmt = f"{r['Linhas']:,}".replace(",", ".")
        md_lines.append(f"| {r['Tabela']} | {linhas_fmt} | {r['Tamanho em disco']} |")

    md_lines += [
        "",
        "## Tabelas vazias (0 linhas)",
        "",
    ]
    vazias = [r["Tabela"] for r in summary_rows if r["Linhas"] == 0]
    md_lines.append(", ".join(f"`{t}`" for t in vazias))
    md_lines.append("")

    with open(output_md, "w", encoding="utf-8") as f:
        f.write("\n".join(md_lines))

    print(f"\nConcluido!")
    print(f"  {output_xlsx}  — Excel com abas Resumo e Colunas")
    print(f"  {output_md}  — Markdown para documentacao")


if __name__ == "__main__":
    main()
